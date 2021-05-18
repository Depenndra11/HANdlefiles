#real column and with for loop read data
import openpyxl

def getRowCount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return (sheet.max_row)


def getColumnCount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return (sheet.max_column)



def readData(file, sheetname, rowNum, ColNum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(row=rowNum, column=ColNum).value



def writeData(file, sheetname, rowNum, ColNum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row=rowNum, column=ColNum).value = data
    workbook.save(file)


